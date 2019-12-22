# -*- coding: utf-8 -*-
"""Преобразование данных из input_file.json в out.xlsx"""


from collections import OrderedDict
import json
import datetime as format_date
import openpyxl as xlsx
from openpyxl.styles.fonts import Font
from openpyxl.styles.borders import Border, Side
from openpyxl.styles.alignment import Alignment


# pylint: disable=too-few-public-methods, simplifiable-if-expression, not-an-iterable
def load_data(json_file) -> dict:
    """Загрузка данных из input_file.json"""
    if json_file is not None:
        with open(json_file, 'r') as file:
            return json.load(file)


MONTHS = {"января": 1, "февраля": 2, "марта": 3, "апреля": 4, "мая": 5, "июня": 6,
          "июля": 7, "августа": 8, "сентября": 9, "октября": 10, "ноября": 11, "декабря": 12}


def sort_data(json_object):
    """Сортировка по дате"""
    list_data = list(json_object.items())
    sort_by_month = lambda element: MONTHS[element[0].split(' ')[1]]
    list_data.sort()
    list_data.sort(key=sort_by_month)
    return OrderedDict(list_data)


def to_date(date):
    """Приведение к нужному формату даты"""
    day, month = date.split(' ')
    year = format_date.datetime.today().year
    return format_date.date(year, MONTHS[month], int(day))


NAME_INDEX = 0
GROUPS_INDEX = 1
CLASSROOM_INDEX = 2
IS_COMPUTER_INDEX = 3


def parse_teachers(list_info):
    """Получение списка ФИО преподователей + параметров по их парам"""
    teachers = []
    if list_info:
        for info in list_info:
            if info:
                name = info[NAME_INDEX]
                attributes = [info[GROUPS_INDEX], info[CLASSROOM_INDEX], info[IS_COMPUTER_INDEX]]
                teachers.append(Teacher(name, attributes))
    return teachers or None


def parse_notes(list_notes: dict, kvant=False) -> tuple:
    """Заметки к списку преподователей"""
    kvants = set()
    notes = [[], []]
    for title, value in list_notes.items():
        teachers = parse_teachers(value)
        if teachers:
            if kvant:
                for teacher in teachers:
                    if teacher.is_kvant:
                        kvants.add(teacher)
                        teachers.remove(teacher)
            if teachers:
                notes[0].append(Note(title, teachers))
    if kvants:
        notes[1].append(Note("K", kvants))
    return tuple(notes) if notes[0] or notes[1] else None


def parse_json(json_object: dict, kvant=False, date_filter=None) -> tuple:
    """Получение списка дней из input_file.json"""
    days = []
    if date_filter:
        date_start, date_end = date_filter.split(':')
        date_start = to_date(date_start)
        date_end = to_date(date_end)
    for date, value in json_object.items():
        date_object = to_date(date)
        if date_filter:
            if date_object < date_start:
                continue
            if date_object > date_end:
                break
        notes = parse_notes(value, kvant)
        if notes:
            day = Day(date, notes)
            days.append(day)
    return tuple(days) or None


class Day:
    """Информация об учебном дне"""
    def __init__(self, date, notes):
        self.date = date
        self.notes = notes[0]
        self.kvants = notes[1]


class Note:
    """Название предмета и список преподователей"""
    def __init__(self, title, teachers):
        self.title = title
        self.teachers = teachers


class Teacher:
    """Информация об преподователе + параметры"""

    def __init__(self, name, attributes):
        self.name = name
        self.groups = attributes[0]
        self.classroom = attributes[1]
        self.is_computer = attributes[2]
        self.is_kvant = True if 'КВАНТ' in self.classroom else False


COLUMNS = {
    'title': 1,
    'teacher': 2,
    'groups': 3,
    'class': 4
}

SHIFT_ROW = 50


def border(style):
    """Границы в таблице"""
    return Border(left=Side(style=style),
                  right=Side(style=style),
                  top=Side(style=style),
                  bottom=Side(style=style))


class ExcelWriter:
    """Работа с Workbook()"""
    def __init__(self):
        self.cur_row = 1
        self.len_title = 0
        self.len_groups = 0
        self.len_teacher = 0
        self.len_classroom = 0
        self.work_book = xlsx.Workbook()


def write(days):
    """Запись в output_file.xlsx"""
    parse_to_xlsx = ExcelWriter()
    work_book = parse_to_xlsx.work_book
    sheet = work_book.get_active_sheet()
    # Поля для печати таблицы
    sheet.page_margins.top = 0.764
    sheet.page_margins.bottom = 0.764
    for day in days:
        write_day(day, parse_to_xlsx)
    # Ширина столбцов + выравнивание
    alignment_parameter = 2
    sheet.column_dimensions['A'].width = parse_to_xlsx.len_title + alignment_parameter
    sheet.column_dimensions['B'].width = parse_to_xlsx.len_teacher + alignment_parameter
    sheet.column_dimensions['C'].width = parse_to_xlsx.len_groups + alignment_parameter
    sheet.column_dimensions['D'].width = parse_to_xlsx.len_classroom + alignment_parameter
    file_name = input("Введите название файла или нажмите 'Enter'(out.xlsx): ")
    if file_name == "":
        file_name = "out.xlsx"
    work_book.save(file_name)


def write_day(day, parse_to_xlsx):
    """Отображение дня в таблице"""
    sheet = parse_to_xlsx.work_book.get_active_sheet()
    sum_rows = 0
    for i in day.notes + day.kvants:
        sum_rows += len(i.teachers)
    while parse_to_xlsx.cur_row % SHIFT_ROW == 0 \
            or (parse_to_xlsx.cur_row % SHIFT_ROW) + sum_rows > SHIFT_ROW:
        parse_to_xlsx.cur_row += 1
    sheet.merge_cells(start_column=COLUMNS['title'], end_column=COLUMNS['class'],
                      start_row=parse_to_xlsx.cur_row, end_row=parse_to_xlsx.cur_row)
    # Центровка и выделение даты
    cell = sheet.cell(row=parse_to_xlsx.cur_row, column=COLUMNS['title'])
    cell.value = day.date
    cell.alignment = Alignment(horizontal='center')
    cell.font = Font(bold=True)
    for column in range(COLUMNS['title'], COLUMNS['class'] + 1):
        sheet.cell(row=parse_to_xlsx.cur_row, column=column).border = border('medium')
    parse_to_xlsx.cur_row += 1
    for kvant in day.kvants:
        write_note(kvant, parse_to_xlsx)
    for note in day.notes:
        write_note(note, parse_to_xlsx)


def write_note(note, parse_to_xlsx):
    """Информация об уроках в таблице"""
    sheet = parse_to_xlsx.work_book.get_active_sheet()
    end_row = parse_to_xlsx.cur_row + len(note.teachers) - 1
    sheet.merge_cells(start_row=parse_to_xlsx.cur_row, end_row=end_row,
                      start_column=COLUMNS['title'], end_column=COLUMNS['title'])
    cell = sheet.cell(row=parse_to_xlsx.cur_row, column=COLUMNS['title'])
    cell.value = note.title
    cell.border = border('thin')
    parse_to_xlsx.len_title = max(parse_to_xlsx.len_title, len(cell.value))
    for teacher in note.teachers:
        # Граница для заголовка в каждой ячейке
        sheet.cell(row=parse_to_xlsx.cur_row, column=COLUMNS['title']).border = border('thin')
        write_teacher(teacher, parse_to_xlsx)


def write_teacher(teacher, parse_to_xlsx):
    """Ячейка преподователя + параметры пары"""
    sheet = parse_to_xlsx.work_book.get_active_sheet()
    cell = sheet.cell(row=parse_to_xlsx.cur_row, column=COLUMNS['teacher'])
    cell.value = teacher.name
    cell.border = border('thin')
    parse_to_xlsx.len_teacher = max(parse_to_xlsx.len_teacher, len(cell.value))
    cell = sheet.cell(row=parse_to_xlsx.cur_row, column=COLUMNS['groups'])
    cell.value = 'гр. ' + ', '.join(teacher.groups)
    cell.border = border('thin')
    parse_to_xlsx.len_groups = max(parse_to_xlsx.len_groups, len(cell.value))
    cell = sheet.cell(row=parse_to_xlsx.cur_row, column=COLUMNS['class'])
    cell_val = 'aуд. ' + str(teacher.classroom)
    cell.value = cell_val + ' (комп.)' if teacher.is_computer else cell_val
    cell.border = border('thin')
    parse_to_xlsx.len_classroom = max(parse_to_xlsx.len_classroom, len(cell.value))
    parse_to_xlsx.cur_row += 1
