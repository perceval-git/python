"""This module writes data received from json-object to .xlsx file"""

import openpyxl as xl
from openpyxl.styles.borders import Border, Side
from openpyxl.styles.fonts import Font
from openpyxl.styles.alignment import Alignment

COLUMNS = {
    'title': 1,
    'teacher': 2,
    'groups': 3,
    'class': 4
}

SHIFT_ROW = 50


def border(style):
    """Return a necessary border"""
    return Border(left=Side(style=style),
                  right=Side(style=style),
                  top=Side(style=style),
                  bottom=Side(style=style))


class ExcelWriter:
    def __init__(self):
        self.cur_row = 1
        self.len_title = 0
        self.len_teacher = 0
        self.len_classroom = 0
        self.len_groups = 0
        self.work_book = xl.Workbook()


def write(days, output):
    """Functions writes data in .xlsx file"""
    xl_parser = ExcelWriter()
    work_book = xl_parser.work_book
    sheet = work_book.get_active_sheet()

    # Setup standard margins to print the table without splits
    sheet.page_margins.top = 0.764
    sheet.page_margins.bottom = 0.764

    for day in days:
        write_day(day, xl_parser)

    # Setup the width of columns
    # This parameter is required to align columns in the table
    alignment_parameter = 2
    sheet.column_dimensions['A'].width = xl_parser.len_title + alignment_parameter
    sheet.column_dimensions['B'].width = xl_parser.len_teacher + alignment_parameter
    sheet.column_dimensions['C'].width = xl_parser.len_groups + alignment_parameter
    sheet.column_dimensions['D'].width = xl_parser.len_classroom + alignment_parameter

    work_book.save(output)


def write_day(day, xl_parser):
    """Setup day to table cells"""
    sheet = xl_parser.work_book.get_active_sheet()

    # Count summary rows in the day
    sum_rows = 0
    for it in day.notes + day.kvants:
        sum_rows += len(it.teachers)

    # Shift the pointer to setup the table without splitting days
    while xl_parser.cur_row % SHIFT_ROW == 0 or (xl_parser.cur_row % SHIFT_ROW) + sum_rows > SHIFT_ROW:
        xl_parser.cur_row += 1

    # Merge cells to setup the date
    sheet.merge_cells(start_column=COLUMNS['title'], end_column=COLUMNS['class'],
                      start_row=xl_parser.cur_row, end_row=xl_parser.cur_row)

    # Setup the date of the day at the center and bold it
    cell = sheet.cell(row=xl_parser.cur_row, column=COLUMNS['title'])
    cell.value = day.date
    cell.alignment = Alignment(horizontal='center')
    cell.font = Font(bold=True)

    # Setup title's borders
    for column in range(COLUMNS['title'], COLUMNS['class'] + 1):
        sheet.cell(row=xl_parser.cur_row, column=column).border = border('medium')

    # Iterate rows of the table
    xl_parser.cur_row += 1

    for kvant in day.kvants:
        write_note(kvant, xl_parser)

    for note in day.notes:
        write_note(note, xl_parser)


def write_note(note, xl_parser):
    """Setup note to table cells that contains info about lessons"""
    sheet = xl_parser.work_book.get_active_sheet()

    # Setup end_row value to then merge cells
    e_row = xl_parser.cur_row + len(note.teachers) - 1
    sheet.merge_cells(start_row=xl_parser.cur_row, end_row=e_row,
                      start_column=COLUMNS['title'], end_column=COLUMNS['title'])

    # Setup the title
    cell = sheet.cell(row=xl_parser.cur_row, column=COLUMNS['title'])
    cell.value = note.title
    cell.border = border('thin')
    xl_parser.len_title = max(xl_parser.len_title, len(cell.value))

    # Setup each teacher with attributes into a row
    for teacher in note.teachers:
        # Setup border to title cells (in each cell)
        sheet.cell(row=xl_parser.cur_row, column=COLUMNS['title']).border = border('thin')
        write_teacher(teacher, xl_parser)


def write_teacher(teacher, xl_parser):
    """Setup teacher to cells table that contain info about teacher such as name, groups and classroom"""
    sheet = xl_parser.work_book.get_active_sheet()

    # Setup teacher's name cell
    cell = sheet.cell(row=xl_parser.cur_row, column=COLUMNS['teacher'])
    cell.value = teacher.name
    cell.border = border('thin')
    xl_parser.len_teacher = max(xl_parser.len_teacher, len(cell.value))

    # Setup groups cell
    cell = sheet.cell(row=xl_parser.cur_row, column=COLUMNS['groups'])
    cell.value = 'гр. ' + ', '.join(teacher.groups)
    cell.border = border('thin')
    xl_parser.len_groups = max(xl_parser.len_groups, len(cell.value))

    # Setup classroom cell with is_computer parameter
    cell = sheet.cell(row=xl_parser.cur_row, column=COLUMNS['class'])
    cell_val = 'aуд. ' + str(teacher.classroom)
    cell.value = cell_val + ' (комп.)' if teacher.is_computer else cell_val
    cell.border = border('thin')
    xl_parser.len_classroom = max(xl_parser.len_classroom, len(cell.value))

    # Iteration of teachers
    xl_parser.cur_row += 1
